from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict, Iterable, Tuple

import numpy as np
import pandas as pd


def _ensure_repo_on_path(bundle_root: Path) -> None:
    if str(bundle_root) not in sys.path:
        sys.path.insert(0, str(bundle_root))


def _safe_sheet_name(name: str) -> str:
    name = "".join(ch for ch in str(name) if ch.isalnum() or ch in (" ", "_", "-")).strip()
    if not name:
        name = "Sheet1"
    return name[:31]


def _unique_sheet_name(proposed: str, used: set[str]) -> str:
    base = _safe_sheet_name(proposed)
    name = base
    idx = 2
    while name in used:
        suffix = f"_{idx}"
        name = f"{base[: max(0, 31 - len(suffix))]}{suffix}"
        idx += 1
    used.add(name)
    return name


def save_table_sci(df: pd.DataFrame, out_dir: Path, stem: str, sheet: str) -> Tuple[Path, Path]:
    """
    References:
      - Science author guidelines: tables should be clear, consistent, and readable.
      - Excel styling via openpyxl for publication-friendly formatting.

    Mathematical expression:
      - None (I/O and formatting only).

    Parameter meanings:
      - df: table content.
      - out_dir: output directory for .csv and .xlsx.
      - stem: base filename without extension.
      - sheet: Excel sheet name (will be sanitized to <= 31 chars).
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / f"{stem}.csv"
    xlsx_path = out_dir / f"{stem}.xlsx"

    df.to_csv(csv_path, index=False)

    sheet_name = _safe_sheet_name(sheet)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"

        from openpyxl.styles import Alignment, Font

        header_font = Font(name="Times New Roman", bold=True)
        body_font = Font(name="Times New Roman", bold=False)
        header_align = Alignment(wrap_text=False, vertical="center")
        body_align = Alignment(wrap_text=False, vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = body_align

        for col in ws.columns:
            max_len = 0
            for cell in col:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 48)

    return csv_path, xlsx_path


def _validate_results_sci_format(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, object]]:
    """
    References:
      - Reproducible research tables: explicit identifiers per row.

    Mathematical expression:
      - Forward-fill grouped identifiers: Chemical_i = Chemical_{i-1} if missing.

    Parameter meanings:
      - df: raw results_sci_format table.
      - Returns: (tidy_df, report) where report includes basic QA checks.
    """
    report: Dict[str, object] = {}
    df_in = df.copy()
    required = [
        "Chemical",
        "Cluster",
        "Place",
        "Average Concentration",
        "CRPS",
        "R2 Total",
        "R2 WLS",
        "R2 Mean",
        "Concentration Std Dev",
        "R2 Std Dev",
        "Percentage (%)",
        "Q0",
        "Pvalue Q0",
        "a",
        "Pvalue a",
        "v0",
        "Pvalue v0",
        "k",
        "Pvalue k",
        "sigma0",
        "Pvalue sigma0",
    ]
    missing = [c for c in required if c not in df_in.columns]
    report["missing_columns"] = missing

    tidy = df_in.copy()
    if "Chemical" in tidy.columns:
        tidy["Chemical"] = tidy["Chemical"].ffill()
    if "Cluster" in tidy.columns:
        tidy["Cluster"] = tidy["Cluster"].ffill()

    if "Place" in tidy.columns:
        places = tidy["Place"].dropna().astype(str).unique().tolist()
        report["places"] = places
        report["invalid_place_values"] = sorted({p for p in places if p not in {"CM", "JH"}})

    report["n_rows"] = int(tidy.shape[0])
    report["n_missing_chemical_raw"] = int(df_in["Chemical"].isna().sum()) if "Chemical" in df_in.columns else None
    report["n_missing_chemical_tidy"] = int(tidy["Chemical"].isna().sum()) if "Chemical" in tidy.columns else None

    pval_cols = [c for c in tidy.columns if str(c).lower().startswith("pvalue")]
    pval_issues = {}
    for col in pval_cols:
        s = pd.to_numeric(tidy[col], errors="coerce")
        bad = int(((s < 0) | (s > 1)).sum())
        if bad:
            pval_issues[col] = bad
    report["pvalue_out_of_range_counts"] = pval_issues
    return tidy, report


def _compute_linear_effect_table(df_sde: pd.DataFrame, cat1_params: pd.DataFrame) -> pd.DataFrame:
    """
    References:
      - Paper figure: SOA_linear_model_benchmark.png (panels g/h use coef * mean, panel i uses coef).
      - Weighted least squares model outputs stored in results_linear_params.

    Mathematical expression:
      - For Model 1: Effect_i = beta_i * mean(env_i * BVOCs_obs)
      - For Model 2: Effect_i = beta_i * mean(env_i * BVOCs_mu_hat)
      - For Model 3: Coef reported as k_temp on C_T_hat (effect panels do not apply).

    Parameter meanings:
      - beta_i: fitted coefficient for predictor i.
      - env_i: environmental proxy term (e.g., RH*NOx, O3*hv).
      - BVOCs_obs: observed BVOCs proxy.
      - BVOCs_mu_hat: temperature-driven expected BVOCs proxy.
    """
    from src.workflow.modeling_framework import _build_env_features

    rows = []
    params = cat1_params.copy()
    params["ModelID"] = params["ModelID"].astype(str)
    for place in sorted(df_sde["place"].dropna().astype(str).unique()):
        df_place = df_sde[df_sde["place"] == place].copy()
        if df_place.empty:
            continue
        env = _build_env_features(df_place)
        for model_id in ["1", "2", "3"]:
            sub_params = params[(params["Place"] == place) & (params["ModelID"] == model_id)].copy()
            if sub_params.empty:
                continue
            for _, r in sub_params.iterrows():
                param = str(r["Parameter"])
                coef = float(r["Estimate"])
                se = float(r["StdErr"]) if pd.notna(r.get("StdErr")) else np.nan
                pval = float(r["p_value"]) if pd.notna(r.get("p_value")) else np.nan
                sig = str(r.get("Significance", "")) if pd.notna(r.get("Significance")) else ""
                if model_id == "3":
                    if param != "C_T_hat":
                        continue
                    reg_mean = float(pd.to_numeric(df_place["bvoc_mu_hat"], errors="coerce").mean())
                else:
                    if param not in env.columns:
                        continue
                    if model_id == "2":
                        bv = pd.to_numeric(df_place["bvoc_mu_hat"], errors="coerce")
                    else:
                        bv = pd.to_numeric(df_place["bvocs"], errors="coerce")
                    reg = pd.to_numeric(env[param], errors="coerce") * bv
                    reg_mean = float(reg.mean())
                rows.append(
                    {
                        "Category": "I",
                        "ModelID": model_id,
                        "Place": place,
                        "Parameter": param,
                        "Estimate": coef,
                        "StdErr": se,
                        "p_value": pval,
                        "Significance": sig,
                        "RegressorMean": reg_mean,
                        "Effect_coef_x_mean": coef * reg_mean,
                        "EffectSE_coef_x_mean": se * reg_mean if np.isfinite(se) else np.nan,
                    }
                )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    model_order = {"1": 1, "2": 2, "3": 3}
    out["ModelOrder"] = out["ModelID"].map(model_order).fillna(99).astype(int)
    out = out.sort_values(["ModelOrder", "Place", "Parameter"]).drop(columns=["ModelOrder"])
    return out.reset_index(drop=True)


def _read_csv_if_exists(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except Exception:
        return pd.DataFrame()


def _build_master_workbook(
    out_path: Path,
    tables: Dict[str, pd.DataFrame],
) -> None:
    """
    References:
      - Science supplementary materials: provide consolidated tables in a single workbook.

    Mathematical expression:
      - None (workbook assembly only).

    Parameter meanings:
      - out_path: output xlsx file path.
      - tables: mapping of sheet_name -> DataFrame.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        index_rows = []
        used_names: set[str] = {"Index"}
        for sheet, df in tables.items():
            sheet_name = _unique_sheet_name(sheet, used_names)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            index_rows.append(
                {
                    "Sheet": sheet_name,
                    "Rows": int(df.shape[0]),
                    "Cols": int(df.shape[1]),
                }
            )
        pd.DataFrame(index_rows).to_excel(writer, sheet_name="Index", index=False)

        from openpyxl.styles import Alignment, Font

        header_font = Font(name="Times New Roman", bold=True)
        body_font = Font(name="Times New Roman", bold=False)
        header_align = Alignment(wrap_text=False, vertical="center")
        body_align = Alignment(wrap_text=False, vertical="center")

        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_align
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = body_font
                    cell.alignment = body_align
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.value is None:
                        continue
                    max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 48)


def parse_args() -> argparse.Namespace:
    bundle_root = Path(__file__).resolve().parents[1]
    repo_root = bundle_root.parent
    parser = argparse.ArgumentParser(description="Export SCI-formatted tables for paper_bundle.")
    parser.add_argument(
        "--results-sci-xlsx",
        type=Path,
        default=repo_root / "results_sci_format.xlsx",
        help="Path to results_sci_format.xlsx generated by VOC fitting workflow.",
    )
    parser.add_argument(
        "--cache-dir",
        type=Path,
        default=bundle_root / "paper" / "checkpoint",
        help="Cache directory containing df_sde.parquet and related caches.",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=bundle_root / "tables",
        help="Output directory (must be paper_bundle\\tables).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    bundle_root = Path(__file__).resolve().parents[1]
    _ensure_repo_on_path(bundle_root)

    out_dir = args.out_dir.resolve()
    cache_dir = args.cache_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    from paper.workflow.lib.modeling_framework_paper import load_cached_results

    df_sde, cat1, cat2, ml, _labels = load_cached_results(cache_dir=cache_dir, tables_dir=out_dir)

    # Core cached tables (Category I/II/III).
    exported = []
    cat1_metrics = cat1.get("metrics", pd.DataFrame())
    cat1_params = cat1.get("params", pd.DataFrame())
    cat2_metrics = cat2.get("metrics", pd.DataFrame()) if isinstance(cat2, dict) else pd.DataFrame()
    cat2_params = cat2.get("params", pd.DataFrame()) if isinstance(cat2, dict) else pd.DataFrame()
    ml_metrics = ml.get("metrics", pd.DataFrame())
    ml_metrics_cs = ml.get("metrics_cs", pd.DataFrame())
    ml_features = ml.get("features", pd.DataFrame())

    if not cat1_metrics.empty:
        exported.append(save_table_sci(cat1_metrics, out_dir, "results_linear_models", "LinearModels"))
    if not cat1_params.empty:
        exported.append(save_table_sci(cat1_params, out_dir, "results_linear_params", "LinearParams"))
    if not cat2_metrics.empty:
        exported.append(save_table_sci(cat2_metrics, out_dir, "results_cs_models", "CSModels"))
    if not cat2_params.empty:
        exported.append(save_table_sci(cat2_params, out_dir, "results_cs_params", "CSParams"))
    if isinstance(ml_metrics, pd.DataFrame) and not ml_metrics.empty:
        exported.append(save_table_sci(ml_metrics, out_dir, "results_ML_models", "MLModels"))
    if isinstance(ml_metrics_cs, pd.DataFrame) and not ml_metrics_cs.empty:
        exported.append(save_table_sci(ml_metrics_cs, out_dir, "results_ML_models_cs", "MLModelsCS"))
    if isinstance(ml_features, pd.DataFrame) and not ml_features.empty:
        exported.append(save_table_sci(ml_features, out_dir, "results_ML_features", "MLFeatures"))

    # SDE parameter table from cache json (if present).
    sde_json = cache_dir / "sde_params.json"
    if sde_json.exists():
        try:
            sde_df = pd.read_json(sde_json)
        except Exception:
            sde_df = pd.DataFrame()
        if not sde_df.empty:
            exported.append(save_table_sci(sde_df, out_dir, "results_sde_params", "SDE"))

    # Linear model effects aligned with SOA_linear_model_benchmark.png.
    linear_effects = _compute_linear_effect_table(df_sde.reset_index(), cat1_params)
    if not linear_effects.empty:
        exported.append(save_table_sci(linear_effects, out_dir, "Table_SI_linear_model_effects", "LinearEffects"))

    # results_sci_format.xlsx validation and export.
    voc_src = args.results_sci_xlsx
    voc_report = {}
    voc_tidy = pd.DataFrame()
    if voc_src.exists():
        voc_raw = pd.read_excel(voc_src, sheet_name=0)
        voc_tidy, voc_report = _validate_results_sci_format(voc_raw)
        # Keep original naming to match workflow expectations.
        exported.append(save_table_sci(voc_raw, out_dir, "results_sci_format", "Results"))
        exported.append(save_table_sci(voc_tidy, out_dir, "Table_SI_results_sci_format_tidy", "VOCFits"))
    else:
        print(f"[WARN] Missing results_sci_format.xlsx at: {voc_src}")

    # Include existing bundle tables if present (do not reformat in-place).
    # Also ensure common "Table_*.csv" outputs have corresponding formatted xlsx.
    table_csv_paths = sorted(out_dir.glob("Table*.csv"))
    for csv_path in table_csv_paths:
        stem = csv_path.stem
        xlsx_path = out_dir / f"{stem}.xlsx"
        if xlsx_path.exists():
            continue
        df_csv = _read_csv_if_exists(csv_path)
        if df_csv.empty:
            continue
        save_table_sci(df_csv, out_dir, stem, stem)

    # Master workbook with key tables as sheets.
    master_tables: Dict[str, pd.DataFrame] = {
        "SDE": sde_df if "sde_df" in locals() and isinstance(sde_df, pd.DataFrame) else pd.DataFrame(),
        "LinearMetrics": cat1_metrics,
        "LinearParams": cat1_params,
        "LinearEffects": linear_effects,
        "CSModels": cat2_metrics,
        "CSParams": cat2_params,
        "MLModels": ml_metrics if isinstance(ml_metrics, pd.DataFrame) else pd.DataFrame(),
        "MLModelsCS": ml_metrics_cs if isinstance(ml_metrics_cs, pd.DataFrame) else pd.DataFrame(),
        "MLFeatures": ml_features if isinstance(ml_features, pd.DataFrame) else pd.DataFrame(),
        "VOCFitsTidy": voc_tidy,
    }

    # Add supplementary tables (skip large bootstrap samples to keep workbook compact).
    exclude_master_stems = {
        "Table_CS_params_bootstrap_samples_CM",
        "Table_CS_params_bootstrap_samples_JH",
    }
    for csv_path in table_csv_paths:
        stem = csv_path.stem
        if stem in exclude_master_stems:
            continue
        df_csv = _read_csv_if_exists(csv_path)
        if isinstance(df_csv, pd.DataFrame) and not df_csv.empty:
            master_tables[stem] = df_csv

    # Drop empty sheets to keep workbook clean.
    master_tables = {k: v for k, v in master_tables.items() if isinstance(v, pd.DataFrame) and not v.empty}
    master_path = out_dir / "Table_SI_All_Tables.xlsx"
    _build_master_workbook(master_path, master_tables)

    # Lightweight QA prints (ASCII only).
    print(f"[OK] Exported tables to: {out_dir}")
    if voc_report:
        print(f"[OK] results_sci_format.xlsx rows={voc_report.get('n_rows')} missing_chemical_raw={voc_report.get('n_missing_chemical_raw')}")
        if voc_report.get("missing_columns"):
            print(f"[WARN] results_sci_format.xlsx missing_columns={voc_report.get('missing_columns')}")
        if voc_report.get("invalid_place_values"):
            print(f"[WARN] results_sci_format.xlsx invalid_place_values={voc_report.get('invalid_place_values')}")
        if voc_report.get("pvalue_out_of_range_counts"):
            print(f"[WARN] results_sci_format.xlsx pvalue_out_of_range={voc_report.get('pvalue_out_of_range_counts')}")
    print(f"[OK] Master workbook: {master_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
